from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Color
from openpyxl.utils import get_column_letter

workbook = Workbook()
worksheet = workbook.active
img = Image.open("backiee-200719.jpg")
img_width = img.size[0]
img_height = img.size[1]
pix = im.load()
for row in range(1, im_height):
    for col in range(1, im_width):
        cell = worksheet.cell(column=col, row=row)
        pixpoint = pix[col - 1, row - 1]
        pixColor = "FF%02X%02X%02X" % (pixpoint[0], pixpoint[1], pixpoint[2])
        fill = PatternFill(patternType='solid', fgColor=Color(rgb=pixColor))
        cell.fill = fill
    worksheet.row_dimensions[row].height = 6
for col in range(1, im_width):
    worksheet.column_dimensions[get_column_letter(col)].width = 1
workbook.save(filename='1049474768.xlsx')
